import csv
from openpyxl import load_workbook, Workbook


wb = load_workbook(filename='input.xlsx')
ws = wb.active

# Create an empty list to store rows
rows = []

# Loop through each row and append to the list
for row in ws.iter_rows():
    rows.append([cell.value for cell in row])

# Save the data to CSV
with open('input.csv', 'w', encoding='utf-8', newline='') as csvfile:
    writer = csv.writer(csvfile, delimiter=',')
    writer.writerows(rows)